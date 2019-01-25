from openpyxl import Workbook
import json

book = Workbook()
sheet = book.active
user = json.load(open('data.json', 'r',encoding='utf-8'))
columns = ['name', 'person', 'line', 'website', 'addr', 'product', 'url', 'phone']

for index, c in enumerate(columns):
    sheet.cell(1, index + 1).value = c
row = 2
for u in user:
    for j, c in enumerate(columns):
        sheet.cell(row, j + 1).value = u[c]
    row = row + 1
book.save(u'company.xlsx')
