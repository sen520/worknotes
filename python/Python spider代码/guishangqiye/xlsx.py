from openpyxl import Workbook
import json

book = Workbook()
sheet = book.active
user = json.load(open('company-info02.json', 'r',encoding='utf-8'))
columns = ['name', 'industry', 'team', 'creator', 'addr1', 'addr2', 'around',
           'type', 'registration']
for index, c in enumerate(columns):
    sheet.cell(1, index + 1).value = c
row = 2
for u in user:
    for j, c in enumerate(columns):
        sheet.cell(row, j + 1).value = u[c]
    row = row + 1
book.save(u'company.xlsx')
