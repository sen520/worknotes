import json, io, uuid, time, re, pymongo



def dataToJson(newDataList, name):
    print(len(newDataList))
    with io.open(name + '.json', 'w', encoding='utf-8') as fo:
        fo.write(json.dumps(newDataList, ensure_ascii=False, indent=2, separators=(',', ': ')))

def json_to_excel(name):
    from openpyxl import Workbook
    import json

    book = Workbook()
    sheet = book.active
    user = json.load(open(name + '.json', 'r', encoding='utf-8'))
    columns = ['name', 'num']

    for index, c in enumerate(columns):
        sheet.cell(1, index + 1).value = c
    row = 2
    for u in user:
        for j, c in enumerate(columns):
            sheet.cell(row, j + 1).value = u[c]
        row = row + 1
    book.save(u'company.xlsx')


if __name__ == '__main__':
    client = pymongo.MongoClient('mongodb://localhost:27017')
    company_names = client.sfda.companyName
    company_name = company_names.find()
    list = []
    for name in company_name:
        nameStr = re.match('(.*) \((.*)\)', name['name'])
        print(nameStr.group(1))
        print(nameStr.group(2))
        list.append({"id": name['id'], 'name': nameStr.group(1), 'num':nameStr.group(2) })
    dataToJson(list, 'companyName')
    json_to_excel('companyName')